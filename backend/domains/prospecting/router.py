"""
Leads API router — zoeken, verrijken, exporteren.

Endpoints:
  POST /api/leads/search              Enkelvoudige SSE-zoekactie
  POST /api/leads/linkedin-people     LinkedIn-profielen zoeken (site:linkedin.com/in)
  POST /api/leads/batch               Batch-zoekactie (template + regio) via SSE
  POST /api/leads/{id}/enrich         Herverrijking (scrape + AI + Hunter)
  POST /api/leads/{id}/hunter         Expliciete Hunter.io-verrijking (domein-zoek + verify)
  POST /api/leads/{id}/waterfall      Waterfall-verrijking (GetLeads → Apollo + Lead Magic telefoon)
  POST /api/leads/{id}/outreach       Start outreach-kwaliteitslus (→ status contacted)
  POST /api/leads/{id}/outreach-send  Genereer + verstuur outreach-mail via SMTP
  GET  /api/leads/stats         Statistieken
  GET  /api/leads/export        Excel-download (volledig NAW)
  GET  /api/leads               Lijst (filter: status, lead_type)
  PATCH /api/leads/{id}         Status updaten
  DELETE /api/leads/{id}        Verwijderen

Status-funnel: new → enriched → valid → contacted → replied
"""
import io
import json
import asyncio
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from . import validate
from .service import LeadsService, BATCH_TEMPLATES, TEMPLATE_LEAD_TYPE

log = logging.getLogger(__name__)
from ...shared.models import LeadUpdate

router = APIRouter(prefix="/api/leads", tags=["leads"])
_svc = LeadsService()


# ── Schemas ───────────────────────────────────────────────────────────────────

class LeadSearchRequest(BaseModel):
    query: str
    max_results: int = 6
    lead_type: str = "overig"
    include_linkedin: bool = False   # LinkedIn-resultaten meenemen


class LinkedInPeopleRequest(BaseModel):
    query: str                       # bijv. "AI directeur zorg Amsterdam"
    max_results: int = 6


class BatchSearchRequest(BaseModel):
    template: str = "custom"         # notarissen_nl | uitvaart_nl | zorg_nl | weareimpact_ai | custom
    queries: List[str] = []          # eigen queries (bij template=custom)
    regio: str = ""                  # optioneel suffix, bijv. "Noord-Holland"
    max_per_query: int = 5
    lead_type: Optional[str] = None  # overschrijft template-default


# ── SSE helper ────────────────────────────────────────────────────────────────

def _sse(data: dict) -> str:
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


# ── Zoeken (enkelvoudig) ──────────────────────────────────────────────────────

@router.post("/search")
async def search_leads(body: LeadSearchRequest):
    async def generate():
        loop = asyncio.get_event_loop()

        yield _sse({"type": "start", "message": f"Zoeken naar '{body.query}'…"})

        results = await loop.run_in_executor(
            None, lambda: _svc.search_web(body.query, body.max_results, include_linkedin=body.include_linkedin)
        )

        if not results:
            yield _sse({"type": "error", "message": "Geen resultaten gevonden"})
            yield _sse({"type": "done"})
            return

        # Filter duplicaten
        results = [r for r in results if not _svc.is_duplicate(r["url"])]
        if not results:
            yield _sse({"type": "done", "message": "Alle gevonden resultaten zijn al in je database."})
            return

        # Zeef artikelen, vacatures en portals eruit vóór het scrapen en de
        # LLM-analyse — zie prospecting/validate.py. Dezelfde zeef als in
        # run_search_batch, zodat de Leads-tab en Iris' lead_search_run dezelfde
        # voorraad opleveren.
        gezeefd, overgeslagen = [], []
        for r in results:
            geschikt, reden = validate.looks_like_organisation(
                r.get("title", ""), r.get("url", ""), r.get("snippet", ""))
            (gezeefd if geschikt else overgeslagen).append(
                r if geschikt else {"titel": r.get("title", "")[:70], "reden": reden})
        results = gezeefd
        if overgeslagen:
            yield _sse({"type": "filtered", "count": len(overgeslagen),
                        "items": overgeslagen[:10]})
        if not results:
            yield _sse({"type": "done",
                        "message": "Geen van de resultaten was een organisatie "
                                   "(alleen artikelen, vacatures of portals)."})
            return

        yield _sse({"type": "found", "count": len(results)})

        for i, r in enumerate(results):
            org_name = validate.clean_org_name(r["title"], r["url"])
            yield _sse({
                "type": "analyzing",
                "org": org_name,
                "index": i + 1,
                "total": len(results),
                "phase": "scrapen",
            })

            scraped = await loop.run_in_executor(
                None, lambda r=r, o=org_name: _svc.scrape_and_enrich(r["url"], o)
            )

            yield _sse({
                "type": "analyzing",
                "org": org_name,
                "index": i + 1,
                "total": len(results),
                "phase": "ai-analyse",
            })

            analysis = await loop.run_in_executor(
                None, lambda r=r, s=scraped, o=org_name: _svc.analyze_lead(
                    o, r["url"], r["snippet"], s
                )
            )

            lead_data = {
                "org_name":    org_name,
                "website":     r["url"],
                "summary":     analysis.get("summary", ""),
                "contacts":    analysis.get("contacts", []),
                "relevance":   analysis.get("relevance", "gemiddeld"),
                "tags":        analysis.get("tags", []),
                "status":      "new",
                "search_query": body.query,
                "lead_type":   body.lead_type,
                "phone":       scraped.get("phone") or analysis.get("phone", ""),
                "email":       scraped.get("email") or analysis.get("email", ""),
                "address":     scraped.get("address") or scraped.get("address_raw", "") or analysis.get("address", ""),
                "city":        scraped.get("city") or analysis.get("city", ""),
                "postal_code": scraped.get("postal_code") or analysis.get("postal_code", ""),
                "kvk_number":  scraped.get("kvk_number") or analysis.get("kvk_number", ""),
                "enriched_at": "",
                "score":       analysis.get("score", 50),
            }

            # Markeer als enriched als we echte NAW-data hebben
            has_naw = bool(lead_data["phone"] or lead_data["address"] or lead_data["email"])
            if has_naw:
                from .service import _now
                lead_data["enriched_at"] = _now()
                lead_data["status"] = "enriched"

            obs_path = await loop.run_in_executor(
                None, lambda d=lead_data: _svc.save_to_obsidian(d)
            )
            lead_data["obsidian_path"] = obs_path or ""

            saved = await loop.run_in_executor(
                None, lambda d=lead_data: _svc.save_to_db(d)
            )

            yield _sse({"type": "lead_saved", "lead": saved})

        yield _sse({"type": "done"})

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Batch-zoekactie ───────────────────────────────────────────────────────────

@router.post("/batch")
async def batch_search(body: BatchSearchRequest):
    """
    Doorloopt een reeks queries (template of eigen) en slaat alle gevonden leads op.
    Streamt voortgang via SSE.
    """
    # Stel queries samen
    if body.template in BATCH_TEMPLATES and not body.queries:
        queries = list(BATCH_TEMPLATES[body.template])
    else:
        queries = list(body.queries)

    if not queries:
        raise HTTPException(status_code=400, detail="Geen queries opgegeven")

    # Voeg regio-suffix toe
    if body.regio:
        queries = [
            (q if body.regio.lower() in q.lower() else f"{q} {body.regio}")
            for q in queries
        ]

    lead_type = body.lead_type or TEMPLATE_LEAD_TYPE.get(body.template, "overig")
    total_queries = len(queries)
    total_saved = 0

    async def generate():
        nonlocal total_saved
        loop = asyncio.get_event_loop()

        yield _sse({
            "type": "batch_start",
            "total_queries": total_queries,
            "template": body.template,
            "lead_type": lead_type,
        })

        for qi, query in enumerate(queries):
            yield _sse({
                "type": "query_start",
                "query": query,
                "query_index": qi + 1,
                "total_queries": total_queries,
            })

            results = await loop.run_in_executor(
                None, lambda q=query: _svc.search_web(q, body.max_per_query)
            )

            # Dedupliceer, en zeef daarna alles wat geen organisatie is
            # (artikel/vacature/portal) — zie prospecting/validate.py.
            new_results = [r for r in results if not _svc.is_duplicate(r["url"])]
            skipped = len(results) - len(new_results)
            gezeefd = []
            for r in new_results:
                geschikt, reden = validate.looks_like_organisation(
                    r.get("title", ""), r.get("url", ""), r.get("snippet", ""))
                if geschikt:
                    gezeefd.append(r)
                else:
                    skipped += 1
                    log.info("[leads] Overgeslagen (%s): %s", reden,
                             (r.get("title") or "")[:70])
            new_results = gezeefd

            if not new_results:
                yield _sse({
                    "type": "query_done",
                    "query": query,
                    "found": 0,
                    "skipped": skipped,
                    "query_index": qi + 1,
                    "total_queries": total_queries,
                })
                continue

            for r in new_results:
                org_name = validate.clean_org_name(r["title"], r["url"])
                yield _sse({
                    "type": "analyzing",
                    "org": org_name,
                    "query_index": qi + 1,
                    "total_queries": total_queries,
                    "phase": "scrapen",
                })

                scraped = await loop.run_in_executor(
                    None, lambda r=r, o=org_name: _svc.scrape_and_enrich(r["url"], o)
                )

                yield _sse({
                    "type": "analyzing",
                    "org": org_name,
                    "query_index": qi + 1,
                    "total_queries": total_queries,
                    "phase": "ai-analyse",
                })

                analysis = await loop.run_in_executor(
                    None, lambda r=r, s=scraped, o=org_name: _svc.analyze_lead(
                        o, r["url"], r["snippet"], s
                    )
                )

                has_naw = bool(
                    scraped.get("phone") or scraped.get("address") or scraped.get("email")
                    or analysis.get("phone") or analysis.get("address") or analysis.get("email")
                )

                from .service import _now
                lead_data = {
                    "org_name":    r["title"],
                    "website":     r["url"],
                    "summary":     analysis.get("summary", ""),
                    "contacts":    analysis.get("contacts", []),
                    "relevance":   analysis.get("relevance", "gemiddeld"),
                    "tags":        analysis.get("tags", []),
                    "status":      "enriched" if has_naw else "new",
                    "search_query": query,
                    "lead_type":   lead_type,
                    "phone":       scraped.get("phone") or analysis.get("phone", ""),
                    "email":       scraped.get("email") or analysis.get("email", ""),
                    "address":     (scraped.get("address") or scraped.get("address_raw", "") or analysis.get("address", "")),
                    "city":        scraped.get("city") or analysis.get("city", ""),
                    "postal_code": scraped.get("postal_code") or analysis.get("postal_code", ""),
                    "kvk_number":  scraped.get("kvk_number") or analysis.get("kvk_number", ""),
                    "enriched_at": _now() if has_naw else "",
                    "score":       analysis.get("score", 80 if has_naw else 50),
                    "obsidian_path": "",
                }

                obs_path = await loop.run_in_executor(
                    None, lambda d=lead_data: _svc.save_to_obsidian(d)
                )
                lead_data["obsidian_path"] = obs_path or ""

                saved = await loop.run_in_executor(
                    None, lambda d=lead_data: _svc.save_to_db(d)
                )

                total_saved += 1
                yield _sse({
                    "type": "lead_saved",
                    "lead": saved,
                    "total_saved": total_saved,
                })

            yield _sse({
                "type": "query_done",
                "query": query,
                "found": len(new_results),
                "skipped": skipped,
                "query_index": qi + 1,
                "total_queries": total_queries,
            })

        yield _sse({
            "type": "batch_done",
            "total_saved": total_saved,
            "total_queries": total_queries,
        })

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── LinkedIn Personen zoeken ───────────────────────────────────────────────────

@router.post("/linkedin-people")
async def search_linkedin_people(body: LinkedInPeopleRequest):
    """
    Zoek LinkedIn-profielen van beslissers in een sector/organisatie.
    Gebruikt site:linkedin.com/in search via Tavily.
    Retourneert resultaten (wordt niet opgeslagen als lead — handmatig selecteren).
    """
    loop = asyncio.get_event_loop()
    results = await loop.run_in_executor(
        None, lambda: _svc.search_linkedin_people(body.query, body.max_results)
    )
    return {"type": "linkedin_people", "query": body.query, "results": results}


# ── Verrijking ────────────────────────────────────────────────────────────────

@router.post("/{lead_id}/enrich")
async def enrich_lead(lead_id: str):
    """Herverrijkt een bestaande lead: scrape + AI + automatisch Hunter als geen e-mail."""
    loop = asyncio.get_event_loop()
    updated = await loop.run_in_executor(
        None, lambda: _svc.enrich_lead(lead_id)
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    return updated


@router.post("/{lead_id}/hunter")
async def hunter_enrich(lead_id: str):
    """
    Expliciete Hunter.io-verrijking: domein-zoeken + e-mailverificatie.
    Zet status op 'valid' als een deliverable e-mail gevonden wordt.
    Vereist HUNTER_API_KEY in .env.
    """
    from ...shared.config import HUNTER_API_KEY
    if not HUNTER_API_KEY:
        raise HTTPException(
            status_code=422,
            detail="HUNTER_API_KEY niet geconfigureerd. Voeg deze toe aan .env en herstart."
        )
    loop = asyncio.get_event_loop()
    updated = await loop.run_in_executor(
        None, lambda: _svc.hunter_enrich(lead_id)
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    return updated


class WaterfallRequest(BaseModel):
    include_phone: bool = True   # ook Lead Magic (telefoon) proberen


@router.post("/{lead_id}/waterfall")
async def waterfall_enrich(lead_id: str, body: WaterfallRequest = WaterfallRequest()):
    """Waterfall-verrijking: GetLeads → Apollo (e-mail) + Lead Magic (telefoon).

    Loopt de goedkopere/nauwkeurigere keten als Hunter geen e-mail opleverde.
    Elke provider is KEY-GATED: zonder key in .env slaat die provider over
    (zie waterfall_report). Geen enkele key → leeg rapport, geen fout.
    """
    loop = asyncio.get_event_loop()
    updated = await loop.run_in_executor(
        None, lambda: _svc.waterfall_enrich(lead_id, include_phone=body.include_phone)
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    return updated


@router.get("/waterfall-status")
def waterfall_status():
    """Welke waterfall-providers zijn geconfigureerd (voor de UI)."""
    from .waterfall import waterfall_report
    return waterfall_report()


# ── Outreach (kwaliteitslus) ──────────────────────────────────────────────────

@router.post("/{lead_id}/outreach", status_code=201)
async def start_lead_outreach(lead_id: str):
    """Start een outreach-kwaliteitslus voor deze lead.

    Bouwt een opdracht uit de gescrapede NAW-/profieldata en zet die in Loop
    Engineering met het Outreach Copywriter (maker) + Outreach Beoordelaar-paar.
    Keert direct terug met loop_id; de frontend volgt 'm in de Loop-tab.
    """
    from ...domains.loop import service as loop_service

    lead = _svc.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")

    contacts = lead.get("contacts") or []
    contact_line = ""
    if contacts:
        c = contacts[0]
        naam = c.get("naam") or ""
        rol = c.get("rol") or ""
        contact_line = f"Contactpersoon: {naam}{(' (' + rol + ')') if rol else ''}\n"

    objective = (
        "Schrijf een warm, persoonlijk B2B-outreachconcept (met onderwerpregel) voor:\n"
        f"Organisatie: {lead.get('org_name', '')}\n"
        f"Type/sector: {lead.get('lead_type', 'overig')}\n"
        f"Plaats: {lead.get('city', '') or 'onbekend'}\n"
        f"Website: {lead.get('website', '') or 'onbekend'}\n"
        f"{contact_line}"
        f"Wat we over hen weten: {lead.get('summary', '') or '—'}\n\n"
        "Aanleiding: relevante samenwerking rond ons herinnerings-/keepsake-aanbod. "
        "Doel: een laagdrempelige kennismaking. Houd het zakelijk, oprecht en AVG-veilig "
        "(alleen zakelijke gegevens, geen consumentenbenadering)."
    )

    maker_id = loop_service.find_profile_id("Outreach Copywriter")
    reviewer_id = loop_service.find_profile_id("Outreach Beoordelaar")
    result = loop_service.spawn_loop(
        objective=objective,
        maker_profile_id=maker_id,
        reviewer_profile_id=reviewer_id,
        threshold=82,
        max_iterations=3,
        session_id=f"outreach-{lead_id[:8]}",
    )

    # Zet status op 'contacted' — outreach-lus is gestart
    from .service import _now
    from ...shared.database import get_conn
    with get_conn() as conn:
        conn.execute(
            "UPDATE leads SET status='contacted', updated_at=? WHERE id=? AND status NOT IN ('replied')",
            (_now(), lead_id),
        )

    return {**result, "lead_id": lead_id, "org_name": lead.get("org_name", "")}


class OutreachSendRequest(BaseModel):
    custom_message: str = ""


class OutreachApproveRequest(BaseModel):
    subject: str = ""   # optioneel: door Vincent aangepaste onderwerpregel
    body: str = ""      # optioneel: door Vincent aangepaste mailtekst


# ── Acquisitie-formule: batch, review-gate en funnel ──────────────────────────

@router.post("/outreach-batch")
async def run_outreach_batch(count: int = Query(0, ge=0, le=50)):
    """Zet nu een batch outreach-concepten klaar ter review (default: dagtarget).

    Verstuurt niets — concepten verschijnen in het Actiecentrum voor goedkeuring."""
    from . import outreach
    return await outreach.prepare_outreach_batch(count)


@router.post("/cleanup-unmailable")
def cleanup_unmailable():
    """Funnel-opschoning: new/enriched leads zonder bruikbaar e-mailadres → lost.
    Verstuurt en verwijdert niets; maakt de voorraadcijfers weer eerlijk."""
    from . import outreach
    return outreach.cleanup_unmailable_leads()


@router.get("/outreach-review")
def list_outreach_review():
    """Alle outreach-concepten die op menselijke goedkeuring wachten."""
    from . import outreach
    leads = _svc.list_leads(status="outreach_review")
    for lead in leads:
        lead["target_email"] = outreach.target_email_for(lead)
    return leads


@router.post("/{lead_id}/outreach-approve")
async def approve_outreach(lead_id: str, body: OutreachApproveRequest = OutreachApproveRequest()):
    """DE verzendknop: verstuur het goedgekeurde concept via Outlook/Graph.

    Dit is de enige plek waar outreach daadwerkelijk de deur uitgaat.
    Status → contacted (met tijdstempel: de input telt mee in de formule)."""
    from . import funnel, outreach
    from ..outlook import service as outlook

    lead = _svc.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")

    subject = (body.subject or lead.get("outreach_subject") or "").strip()
    mail_body = (body.body or lead.get("outreach_draft") or "").strip()
    if not subject or not mail_body:
        raise HTTPException(status_code=422, detail="Geen concept aanwezig — draai eerst de outreach-batch.")

    target = outreach.target_email_for(lead)
    if not target:
        raise HTTPException(status_code=422, detail="Geen e-mailadres bekend voor deze lead.")
    ok, why = outreach.valid_target(lead)
    if not ok:
        raise HTTPException(
            status_code=422,
            detail=f"Dit adres ({target}) is geen serieus prospect-adres ({why}) — versturen geweigerd. "
                   "Zoek een specifiek contact of wijs de lead af.",
        )
    # GDPR-opt-out-guard: staat het adres (of zijn domein) op de blocklist,
    # dan gaan we hard op slot — iemand die 'STOP' zei mag nooit herdreigen.
    from . import opt_out
    if opt_out.is_opted_out(target):
        raise HTTPException(
            status_code=422,
            detail=f"{target} staat op de opt-out-blocklist (afgemeld). Versturen geweigerd — "
                   "dat is wettelijk verplicht.",
        )
    # Echte token-check (niet alleen de gecachte account): een verlopen
    # refresh-token laat is_authenticated() op True staan maar geeft géén
    # token -> dat zou anders een ongevangen RuntimeError (HTTP 500) geven.
    if not outlook.get_valid_token():
        raise HTTPException(
            status_code=422,
            detail="Outlook-sessie ongeldig of verlopen — log opnieuw in via "
                   "Instellingen → Outlook en probeer het daarna opnieuw.",
        )

    result = await outlook.send_new_email(
        to=target, subject=subject, body_html=mail_body.replace("\n", "<br>"),
    )
    if not result.get("success"):
        raise HTTPException(
            status_code=502,
            detail=f"Versturen mislukt: {result.get('error', result)}",
        )

    # Bewaar wat er echt verstuurd is (evt. door Vincent aangepast) als record.
    from ...shared.database import get_conn
    from .service import _now
    with get_conn() as conn:
        conn.execute(
            "UPDATE leads SET outreach_subject = ?, outreach_draft = ?, updated_at = ? WHERE id = ?",
            (subject, mail_body, _now(), lead_id),
        )
    updated = funnel.advance_lead(lead_id, "contacted")

    from ...shared.outcomes import log_outcome
    log_outcome(
        "Leads", "outreach_sent",
        f"Outreach verstuurd aan {lead['org_name']} ({target}): '{subject}'",
        next_step="Reply-detectie staat aan — je hoort het zodra ze reageren.",
    )
    return {"status": "sent", "to": target, "subject": subject, "lead": updated}


@router.post("/{lead_id}/outreach-dismiss")
def dismiss_outreach(lead_id: str):
    """Wijs een concept af: de lead gaat naar 'lost' (met tijdstempel).

    Zonder deze zijuitgang zou een afgewezen lead de volgende ochtend gewoon
    weer in de batch opduiken. Toch nog benaderen? Zet de status handmatig
    terug via PATCH /api/leads/{id} — dan doet hij weer mee."""
    from . import funnel
    lead = _svc.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    from ...shared.database import get_conn
    from .service import _now
    with get_conn() as conn:
        conn.execute(
            "UPDATE leads SET outreach_subject = '', outreach_draft = '', "
            "outreach_drafted_at = '', updated_at = ? WHERE id = ?",
            (_now(), lead_id),
        )
    funnel.advance_lead(lead_id, "lost")
    return {"status": "dismissed", "lead_id": lead_id, "back_to": "lost"}


# ── Follow-up (zachte herinnering na stilte) ──────────────────────────────

@router.post("/{lead_id}/followup-approve")
async def approve_followup(lead_id: str, body: OutreachApproveRequest = OutreachApproveRequest()):
    """Verstuur het opvolgconcept — dezelfde guards als outreach-approve
    (Outlook-token, opt-out-blocklist, geldig adres), want dit is nog steeds
    ongevraagde commerciële mail en valt onder dezelfde wetgeving."""
    from . import outreach, opt_out
    from . import followup as followup_service
    from ..outlook import service as outlook

    lead = _svc.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")

    subject = (body.subject or lead.get("followup_subject") or "").strip()
    mail_body = (body.body or lead.get("followup_draft") or "").strip()
    if not subject or not mail_body:
        raise HTTPException(status_code=422, detail="Geen opvolgconcept aanwezig.")

    target = outreach.target_email_for(lead)
    if not target:
        raise HTTPException(status_code=422, detail="Geen e-mailadres bekend voor deze lead.")
    ok, why = outreach.valid_target(lead)
    if not ok:
        raise HTTPException(status_code=422, detail=f"Dit adres ({target}) is geen serieus prospect-adres ({why}).")
    if opt_out.is_opted_out(target):
        raise HTTPException(status_code=422, detail=f"{target} staat op de opt-out-blocklist.")
    if not outlook.get_valid_token():
        raise HTTPException(
            status_code=422,
            detail="Outlook-sessie ongeldig of verlopen — log opnieuw in via Instellingen → Outlook.",
        )

    result = await outlook.send_new_email(
        to=target, subject=subject, body_html=mail_body.replace("\n", "<br>"),
    )
    if not result.get("success"):
        raise HTTPException(status_code=502, detail=f"Versturen mislukt: {result.get('error', result)}")

    followup_service.na_verzending(lead_id)
    return {"status": "sent", "to": target, "subject": subject}


@router.post("/{lead_id}/followup-dismiss")
def dismiss_followup(lead_id: str):
    """Bewust geen tweede poging deze ronde — telt mee als afgehandelde
    poging, anders herverschijnt hetzelfde concept morgen weer (zie
    prospecting/followup.py)."""
    from . import followup as followup_service
    lead = _svc.get_lead(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    followup_service.sla_followup_over(lead_id)
    return {"status": "dismissed", "lead_id": lead_id}


# ── Opt-out (Telecommunicatiewet art. 11.7) ───────────────────────────────────

class OptOutRequest(BaseModel):
    email: str
    snippet: str = ""        # optioneel: de reply-tekst waarin 'STOP' stond


@router.post("/opt-out")
def register_opt_out(body: OptOutRequest):
    """Blokkeer een adres (én domein) voor toekomstige outreach.

    Wordt aangeroepen als een reply 'STOP' / 'afmelden' bevat, of handmatig
    vanuit de UI. Returns of er iets nieuws geblokkeerd is."""
    from . import opt_out
    blocked = opt_out.record_opt_out(body.email, source="manual", raw_snippet=body.snippet)
    return {"status": "opted_out" if blocked else "already_blocked", "email": body.email}


@router.get("/opt-out/list")
def list_opt_outs():
    from . import opt_out
    return {"opt_outs": opt_out.list_opt_outs()}


@router.post("/opt-out/check")
def check_opt_out(email: str):
    from . import opt_out
    return {"email": email, "opted_out": opt_out.is_opted_out(email)}


@router.get("/funnel")
def funnel_overview():
    """De conversieformule: funnel-standen, ratio's en geleverde inputs (7 dagen)."""
    from . import funnel
    return {**funnel.funnel_stats(), "inputs": funnel.input_stats(days=7)}


@router.post("/{lead_id}/outreach-send")
async def send_lead_outreach(lead_id: str, body: OutreachSendRequest = OutreachSendRequest()):
    """
    Genereer een persoonlijke outreach-mail via Hermes en verstuur via SMTP.
    De mail wordt verstuurd vanaf v.munster@weareimpact.nl naar het e-mailadres van de lead.
    Status wordt gezet op 'contacted' na succesvol versturen.
    """
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None, lambda: _svc.send_outreach_email(lead_id, body.custom_message)
    )
    return result


# ── Statistieken ──────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats():
    return _svc.get_stats()


# ── Quality Gate (Hermes Lead Machine: "gooi de mismatches weg") ───────────────

@router.post("/quality-gate")
def quality_gate(dry_run: bool = False):
    """Beoordeel elke onbenaderde lead op fit en zet mismatches naar 'lost'.

    De automatische versie van de SCORE-stap uit de video: rangschikt 0-100 en
    verwijdert stil de leads die niet passen (status → 'lost', reden
    'quality_gate'). Verstuurt en verwijdert niets; de lead blijft in de DB."""
    from . import quality_gate
    return quality_gate.run_quality_gate(dry_run=dry_run)


@router.get("/quality-gate/summary")
def quality_gate_summary():
    from . import quality_gate
    return quality_gate.quality_summary()


@router.get("/top")
def top_leads(limit: int = Query(10, ge=1, le=50)):
    """De schone lijst voor bovenaan de pagina: scherpe fits die nog wachten
    op een besluit (nog niet gecontacteerd/verloren), hoogste score eerst.
    Dit is het antwoord op 'wat zijn nu mijn beste leads' zonder dat je door
    de volle tabel hoeft te scrollen."""
    from . import quality_gate
    with_status = ("new", "enriched", "valid")
    leads = [
        l for l in _svc.list_leads()
        if l.get("status") in with_status
        and int(l.get("quality_score") or l.get("score") or 0) >= quality_gate.QUALITY_TARGET_SCORE
    ]
    leads.sort(key=lambda l: int(l.get("quality_score") or l.get("score") or 0), reverse=True)
    return leads[:limit]


# ── Beschrijf-in-1-zin → volledige pipeline (video-UX: "type one sentence") ───

class DescribeRequest(BaseModel):
    sentence: str                      # vrije taal, bv. "SEO agencies die linkbuilding doen"
    max_results: int = 5
    lead_type: str = "overig"
    run_quality_gate: bool = True      # gooi mismatches direct weg na verrijking
    run_waterfall: bool = True         # haal telefoon/e-mail op voor score >= 70 (A/B)


@router.post("/describe")
async def describe_run(body: DescribeRequest):
    """De 'type één zin' instap uit de video.

    Beschrijft wie je wil bereiken in gewone taal, zet dat om in een zoekquery,
    draait de volledige find→enrich→verify→score-pipeline en (optioneel) de
    Quality Gate. Streamt voortgang via SSE, net als /search maar met een
    natuurlijke-taal-front. Verstuurt NIETS — dat blijft achter de review-gate.
    """
    async def generate():
        loop = asyncio.get_event_loop()
        yield _sse({"type": "start",
                    "message": f"Pipeline gestart voor: '{body.sentence}'"})

        # 1. DESCRIBE → gestructureerde zoekquery via Hermes (NL zin → query)
        query = await loop.run_in_executor(
            None, lambda: _svc.describe_to_query(body.sentence)
        )
        yield _sse({"type": "describe", "query": query})

        # 2-4. FIND → ENRICH → VERIFY via de bestaande zoekketen
        results = await loop.run_in_executor(
            None, lambda: _svc.search_web(query, body.max_results)
        )
        results = [r for r in results if not _svc.is_duplicate(r["url"])] if results else []
        if not results:
            yield _sse({"type": "done", "message": "Geen organisaties gevonden voor die omschrijving."})
            return

        saved = []
        for r in results:
            geschikt, reden = validate.looks_like_organisation(
                r.get("title", ""), r.get("url", ""), r.get("snippet", ""))
            if not geschikt:
                continue
            org_name = validate.clean_org_name(r["title"], r["url"])
            scraped = await loop.run_in_executor(
                None, lambda r=r, o=org_name: _svc.scrape_and_enrich(r["url"], o))
            analysis = await loop.run_in_executor(
                None, lambda r=r, s=scraped, o=org_name: _svc.analyze_lead(
                    o, r["url"], r["snippet"], s))
            lead_data = {
                "org_name": org_name, "website": r["url"],
                "summary": analysis.get("summary", ""),
                "contacts": analysis.get("contacts", []),
                "relevance": analysis.get("relevance", "gemiddeld"),
                "tags": analysis.get("tags", []),
                "status": "new", "search_query": body.sentence,
                "lead_type": body.lead_type,
                "phone": scraped.get("phone") or analysis.get("phone", ""),
                "email": scraped.get("email") or analysis.get("email", ""),
                "address": scraped.get("address") or scraped.get("address_raw", "") or analysis.get("address", ""),
                "city": scraped.get("city") or analysis.get("city", ""),
                "postal_code": scraped.get("postal_code") or analysis.get("postal_code", ""),
                "kvk_number": scraped.get("kvk_number") or analysis.get("kvk_number", ""),
                "enriched_at": "",
                "score": analysis.get("score", 50),
            }
            has_naw = bool(lead_data["phone"] or lead_data["address"] or lead_data["email"])
            if has_naw:
                from .service import _now
                lead_data["enriched_at"] = _now()
                lead_data["status"] = "enriched"
            obs_path = await loop.run_in_executor(
                None, lambda d=lead_data: _svc.save_to_obsidian(d))
            lead_data["obsidian_path"] = obs_path or ""
            saved_lead = await loop.run_in_executor(
                None, lambda d=lead_data: _svc.save_to_db(d))
            saved.append(saved_lead)
            yield _sse({"type": "lead_saved", "lead": saved_lead})

        # 5. SCORE + (optioneel) discard — de Quality Gate
        if body.run_quality_gate and saved:
            from . import quality_gate
            gate = await loop.run_in_executor(None, lambda: quality_gate.run_quality_gate())
            yield _sse({"type": "quality_gate",
                        "promoted": len(gate["promoted"]),
                        "discarded": len(gate["discarded"]),
                        "kept": len(gate["kept"])})

        # 6. WATERFALL — alleen voor de scherpe fits (score >= 70, A/B), en
        # alleen als de Quality Gate ook echt draaide (anders weten we het
        # label niet en zou dit voor élke gevonden lead gelden, incl. ruis).
        top_ids = []
        if body.run_waterfall and body.run_quality_gate and saved:
            from . import quality_gate
            for s in saved:
                score = s.get("score") or 0
                try:
                    score = int(score)
                except (TypeError, ValueError):
                    score = 0
                if score >= quality_gate.QUALITY_TARGET_SCORE:
                    top_ids.append(s["id"])
            for lead_id in top_ids:
                updated = await loop.run_in_executor(
                    None, lambda lid=lead_id: _svc.waterfall_enrich(lid, include_phone=True))
                if updated:
                    yield _sse({"type": "waterfall_enriched", "lead": updated})

        yield _sse({"type": "done",
                    "message": f"{len(saved)} lead(s) gevonden"
                               f"{f', {len(top_ids)} topfit verrijkt met contactgegevens' if top_ids else ''}. "
                               f"Review ze in het overzicht — nog niets verstuurd."})
    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/export")
def export_leads(
    status: Optional[str] = Query(None),
    lead_type: Optional[str] = Query(None),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    leads = _svc.list_leads(status=status, lead_type=lead_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Bedrijfsnaam", "Type", "Adres", "Postcode", "Stad",
        "Telefoon", "E-mail", "Website", "KvK", "Contactpersoon",
        "Functie", "Contact Email", "Samenvatting",
        "Relevantie", "Status", "Zoekopdracht", "Verrijkt op", "Aangemaakt",
    ]

    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    HDR_FONT  = Font(bold=True, color="E2E8F0", size=10)
    HOOG_FILL = PatternFill("solid", fgColor="052E16")  # groen
    MIDDEN_FILL = PatternFill("solid", fgColor="431407")  # amber
    LAAG_FILL = PatternFill("solid", fgColor="1E293B")   # slate

    COL_WIDTHS = [35, 12, 35, 11, 22, 16, 32, 42, 11, 25, 20, 32, 70, 11, 12, 30, 20, 14]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="left")

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    row_num = 2
    for lead in leads:
        contacts = lead.get("contacts") or []

        rel = lead.get("relevance", "gemiddeld")
        rel_fill = {"hoog": HOOG_FILL, "laag": LAAG_FILL}.get(rel, MIDDEN_FILL)

        def write_row(contact: dict = None):
            nonlocal row_num
            naam = (contact or {}).get("naam", "")
            rol  = (contact or {}).get("rol", "")
            cem  = (contact or {}).get("email", "")
            enriched = (lead.get("enriched_at") or "")[:10]
            ws.append([
                lead.get("org_name", ""),
                lead.get("lead_type", "overig"),
                lead.get("address", ""),
                lead.get("postal_code", ""),
                lead.get("city", ""),
                lead.get("phone", ""),
                lead.get("email", ""),
                lead.get("website", ""),
                lead.get("kvk_number", ""),
                naam, rol, cem,
                lead.get("summary", ""),
                rel,
                lead.get("status", ""),
                lead.get("search_query", ""),
                enriched,
                (lead.get("created_at") or "")[:10],
            ])
            # Kleurcode op relevantie
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = rel_fill
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=False)
            row_num += 1

        if contacts:
            for c in contacts:
                write_row(c)
        else:
            write_row()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("")
def list_leads(
    status: Optional[str] = Query(None),
    lead_type: Optional[str] = Query(None),
):
    return _svc.list_leads(status=status, lead_type=lead_type)


@router.patch("/{lead_id}")
def update_lead(lead_id: str, body: LeadUpdate):
    updated = _svc.update_status(lead_id, body.status)
    if not updated:
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
    return updated


@router.delete("/{lead_id}", status_code=204)
def delete_lead(lead_id: str):
    if not _svc.delete_lead(lead_id):
        raise HTTPException(status_code=404, detail="Lead niet gevonden")
